import serial
import time
import os
from openpyxl import Workbook, load_workbook

# Serial Communication Settings
arduino_port = 'COM3'  # Change as per your setup
baud_rate = 9600


# Excel File Setup
excel_file = r"D:\workvv\arduino\SoilMoistureLog.xlsx"

# Moisture Threshold
MOISTURE_THRESHOLD = 30  

# Initialize Excel File
def initialize_excel():
    """Create and initialize the Excel file if it does not exist."""
    try:
        if os.path.exists(excel_file):
            workbook = load_workbook(excel_file)
            print(f"Loaded existing file: {excel_file}")
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Soil Moisture Logs"
            sheet.append(["Timestamp", "Date", "Time", "Moisture Level (%)", "Flow Rate (mL/min)", "Motor Status"])
            workbook.save(excel_file)
            print(f"Created new file: {excel_file}")
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        exit(1)
    return workbook

# Log Data to Excel
def log_to_excel(timestamp, date, time, moisture_level, flow_rate, motor_status):
    """Log soil moisture, water flow, and motor status to Excel file."""
    try:
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        sheet.append([timestamp, date, time, moisture_level, flow_rate, motor_status])
        workbook.save(excel_file)
        print(f"Logged: {timestamp}, Moisture: {moisture_level}%, Flow Rate: {flow_rate} mL/min, Motor: {motor_status}")
    except PermissionError:
        print("Error: Excel file is open. Please close it and restart the program.")

# Initialize Serial Connection
print("Initializing serial connection...")
try:
    arduino = serial.Serial(arduino_port, baud_rate, timeout=1)
    time.sleep(2)  # Allow Arduino to stabilize
    print("Serial connection established!")
except serial.SerialException as e:
    print(f"Error: Could not open port {arduino_port}. {e}")
    exit(1)

# Initialize Excel file
initialize_excel()

# Main Loop to Read Data
print("Listening for data from Arduino...")

try:
    while True:
        if arduino.in_waiting > 0:
            message = arduino.readline().decode('utf-8', errors='ignore').strip()

            # Ignore startup messages
            if "Smart Irrigation System Initialized" in message or message == "":
                print(f"Skipping invalid data: {message}")
                continue

            # Validate message format
            if "Moisture:" in message and "FlowRate:" in message and "Motor:" in message:
                try:
                    data = message.split(", ")
                    moisture_level = int(data[0].split(":")[1])
                    flow_rate = float(data[1].split(":")[1])
                    motor_status = data[2].split(":")[1]

                    print(f"Moisture: {moisture_level}%, Flow Rate: {flow_rate} mL/min, Motor: {motor_status}")

                    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
                    date = time.strftime("%Y-%m-%d")
                    current_time = time.strftime("%H:%M:%S")

                    # Log data to Excel
                    log_to_excel(timestamp, date, current_time, moisture_level, flow_rate, motor_status)

                except Exception as e:
                    print(f"Error processing data: {e}")

            else:
                print(f"Skipping invalid data: {message}")

        time.sleep(1)  # Polling delay

except KeyboardInterrupt:
    print("Program terminated by user.")

finally:
    arduino.close()
    print("Serial connection closed.")
